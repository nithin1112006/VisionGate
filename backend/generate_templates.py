import openpyxl
import os

templates_dir = r"d:\CLG\AAS\APP\Attenda_V1 - Staff - Copy\templates"
os.makedirs(templates_dir, exist_ok=True)

# 1. Departments Template
wb_dept = openpyxl.Workbook()
ws_dept = wb_dept.active
ws_dept.title = "Departments"
ws_dept.append(["Department Name"])
ws_dept.append(["Computer Science & Engineering"])
ws_dept.append(["Information Technology"])
ws_dept.append(["Artificial Intelligence & Machine Learning"])
ws_dept.append(["Mechanical Engineering"])
ws_dept.append(["Electronics & Communication"])

wb_dept.save(os.path.join(templates_dir, "departments_template.xlsx"))
print("Created departments_template.xlsx")

# 2. Users / Staff Template
wb_users = openpyxl.Workbook()
ws_users = wb_users.active
ws_users.title = "Staff Users"
ws_users.append(["Username", "Name", "Department", "Role", "Password"])
ws_users.append(["alice_smith", "Alice Smith", "Computer Science & Engineering", "hod", "securePassword123"])
ws_users.append(["bob_jones", "Bob Jones", "Information Technology", "staff", "staffPass321"])
ws_users.append(["charlie_brown", "Charlie Brown", "Computer Science & Engineering", "staff", "charliePass!"])

ws_users["G1"] = "Available Roles:"
roles = ["hod", "staff"]
for i, r in enumerate(roles, start=2):
    ws_users.cell(row=i, column=7, value=r)
ws_users.column_dimensions['G'].width = 25

ws_users["H1"] = "Available Departments:"
mock_depts = [
    "Computer Science & Engineering",
    "Information Technology",
    "Artificial Intelligence & Machine Learning",
    "Mechanical Engineering",
    "Electronics & Communication"
]
for i, d in enumerate(mock_depts, start=2):
    ws_users.cell(row=i, column=8, value=d)
ws_users.column_dimensions['H'].width = 35

wb_users.save(os.path.join(templates_dir, "users_template.xlsx"))
print("Created users_template.xlsx")
